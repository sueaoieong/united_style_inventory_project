import sys,us_main,datetime
from in_ui import Ui_Form
from PyQt5.QtWidgets import QDialog,QApplication,QTableWidgetItem,QMessageBox
from PyQt5.QtGui import QIntValidator
from openpyxl import Workbook
from PyQt5 import QtCore

class in_show(QDialog):
    def __init__(self):
        super().__init__()
        self.ui=Ui_Form()
        self.db=us_main.us_main()
        self.barcode_t=[]
        self.barcode_qty=[]
        self.ui.setupUi(self)
        self.ui.lineEdit_4.setValidator(QIntValidator())
        self.table_create()
        self.ui.pushButton.clicked.connect(self.get_info)
        self.ui.pushButton_2.clicked.connect(self.save_data)
        self.show()

    def save_data(self):

        date=datetime.datetime.today().strftime("%Y-%m-%d")
        totalrow=self.ui.tableWidget.rowCount()
        print(totalrow)

        wb=Workbook()
        ws=wb.active
        dest_filename="barcode-list_"+date+".xlsx"
        j=1
        ws.cell(column=1,row=j, value="貨名")
        ws.cell(column=2,row=j, value="條碼")
        if len(self.barcode_t)!=0:
            for i in range(len(self.barcode_t)):
                sql="SELECT 貨名 from inventory where Barcode=?"
                temp=self.db.search_item(self.db.conn,sql,(self.barcode_t[i],))
                print(temp[0])
                name=temp[0]
                for k in range(self.barcode_qty[i]):
                    j+=1
                    ws.cell(column=1,row=j,value=name)
                    ws.cell(column=2,row=j,value=self.barcode_t[i])
        if totalrow!=0:
            for i in range(totalrow):
                list=[]
                for j in range(5):
                    list.append(self.ui.tableWidget.item(i,j).text())
                print(list)
                sql="INSERT INTO in_DB(post_ID, 貨名, 顏色, 尺寸,庫存, 日期) VALUES (?,?,?,?,?,?)"
                value=(list[0],list[1],list[2],list[3],int(list[4]),date)
                self.db.add_item(self.db.conn,sql,value)
            wb.save(dest_filename)
            print(totalrow)
            for i in range(totalrow):
                self.ui.tableWidget.removeRow(0)
            print("=====check barcode=====")
            print(self.barcode_t)
            print(self.barcode_qty)
            sql1="SELECT * from goods_amount where Barcode=?"
            sql2="UPDATE goods_amount SET Qty = ?  where Barcode=?"
            sql3="INSERT INTO goods_amount(Barcode,Qty) VALUES (?,?)"
            for i in range(len(self.barcode_t)):
                #check if barcode exist
                temp=self.db.search_item(self.db.conn,sql1,(self.barcode_t[i],))
                print("======goods_amount======")
                print(temp)

                #get the result of Qty
                #if yes, update item
                #if no, add item
                if temp == None:
                    value=(self.barcode_t[i],self.barcode_qty[i])
                    self.db.add_item(self.db.conn,sql3,value)
                else:
                    B=temp[0]
                    T=temp[1]
                    self.db.update_item(self.db.conn,sql2,(T+self.barcode_qty[i],B))

            msg=QMessageBox()
            msg.setWindowTitle("Information")
            msg.setText("儲存完成!")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

    def table_create(self):
        #rowposition=self.ui.tableWidget.rowCount()
        #self.ui.tableWidget.insertRow(rowposition)

        self.ui.tableWidget.setColumnCount(5)
        self.ui.tableWidget.setHorizontalHeaderLabels(["post ID","貨名","顏色","尺寸","庫存"])
        #self.ui.tableWidget.setItem(0,0,QTableWidgetItem("ad"))
        #self.ui.tableWidget.setItem(0,1,QTableWidgetItem("ad"))
        #self.ui.tableWidget.setItem(0,2,QTableWidgetItem("ad"))

    def get_info(self):
        id=self.ui.lineEdit.text()
        sql="SELECT 貨名 FROM goods where post_ID=?"

        temp=self.db.search_item(self.db.conn,sql,(id,))
        name=temp[0]
        color=self.ui.lineEdit_2.text()
        size=self.ui.lineEdit_3.text()
        qty=self.ui.lineEdit_4.text()

        self.barcode_t.append(id+color+size)
        self.barcode_qty.append(int(qty))
        print("====get_info checking====")
        print(self.barcode_t)
        print(self.barcode_qty)
        print(type(qty),qty)

        row=self.ui.tableWidget.rowCount()
        self.ui.tableWidget.insertRow(row)
        self.ui.tableWidget.setItem(row,0,QTableWidgetItem(id))
        self.ui.tableWidget.setItem(row,1,QTableWidgetItem(name))
        self.ui.tableWidget.setItem(row,2,QTableWidgetItem(color))
        self.ui.tableWidget.setItem(row,3,QTableWidgetItem(size))
        self.ui.tableWidget.setItem(row,4,QTableWidgetItem(qty))


if __name__=="__main__":
    app = QApplication(sys.argv)
    app.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling, True)
    w=in_show()
    sys.exit(app.exec_())
