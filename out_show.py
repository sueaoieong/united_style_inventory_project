import sys,us_main,datetime
from out_ui import Ui_Form
from PyQt5.QtWidgets import QDialog,QApplication,QTableWidgetItem,QMessageBox
from openpyxl import Workbook
from PyQt5 import QtCore

class out_show(QDialog):
    def __init__(self):
        super().__init__()
        self.ui=Ui_Form()
        self.ui.setupUi(self)
        self.db=us_main.us_main()
        self.barcode_t=[]
        self.barcode_qty=[]
        self.fetch_people()
        self.table_create()
        self.ui.lineEdit_2.returnPressed.connect(self.get_info)
        self.ui.pushButton.setAutoDefault(False)
        self.ui.pushButton.clicked.connect(self.save_data)
        self.show()


    def fetch_people(self):
        sql="SELECT 人名 from people"
        result=self.db.search_all(self.db.conn,sql)
        for i in result:
            temp=i[0]

            self.ui.comboBox.addItem(temp)

    def table_create(self):
        self.ui.tableWidget.setColumnCount(4)
        self.ui.tableWidget.setHorizontalHeaderLabels(["Barcode","貨名","顏色","尺寸"])

    def save_data(self):
        date=datetime.datetime.today().strftime("%Y-%m-%d")
        name=self.ui.comboBox.currentText()
        print(name)
        totalrow=self.ui.tableWidget.rowCount()
        print(totalrow)

        wb=Workbook()
        dest_filename=name+"_"+date+".xlsx"
        ws=wb.active

        j=1
        ws.cell(column=1,row=j, value="條碼")
        ws.cell(column=2,row=j, value="貨名")
        ws.cell(column=3,row=j, value="顏色")
        ws.cell(column=4,row=j, value="尺寸")

        if totalrow!=0:
            for i in range(totalrow):
                barcode=self.ui.tableWidget.item(i,0).text()
                goodname=self.ui.tableWidget.item(i,1).text()
                color=self.ui.tableWidget.item(i,2).text()
                size=self.ui.tableWidget.item(i,3).text()
                j+=1
                ws.cell(column=1,row=j, value=barcode)
                ws.cell(column=2,row=j, value=goodname)
                ws.cell(column=3,row=j, value=color)
                ws.cell(column=4,row=j, value=size)
                sql="INSERT INTO out_DB(交收人員, Barcode, 日期) VALUES (?,?,?)"
                value=(name,barcode,date)
                self.db.add_item(self.db.conn,sql,value)
            wb.save(dest_filename)
            print(totalrow)
            for i in range(totalrow):
                self.ui.tableWidget.removeRow(0)
            print("=====check barcode=====")
            print(self.barcode_t)
            print(self.barcode_qty)
            sql1="SELECT * from goods_amount where Barcode=?"
            sql2="UPDATE goods_amount SET Qty = ?  where Barcode=?"
            sql3="INSERT INTO goods_amount(Barcode,Qty) VALUES (?,?)"
            for i in range(len(self.barcode_t)):
                #check if barcode exist
                temp=self.db.search_item(self.db.conn,sql1,(self.barcode_t[i],))
                print("======goods_amount======")
                print(temp)

                #get the result of Qty
                #if yes, update item
                #if no, add item
                if temp == None:
                    value=(self.barcode_t[i],self.barcode_qty[i])
                    self.db.add_item(self.db.conn,sql3,value)
                else:
                    B=temp[0]
                    T=temp[1]
                    self.db.update_item(self.db.conn,sql2,(T-self.barcode_qty[i],B))


            print(self.barcode_t)
            print(self.barcode_qty)

            msg=QMessageBox()
            msg.setWindowTitle("Information")
            msg.setText("儲存完成!")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

    def get_info(self):

        barcode=self.ui.lineEdit_2.text()

        if barcode=="":
            pass
        else:
            k=0
            #print(len(self.barcode_t))
            for i in range(len(self.barcode_t)):
                if self.barcode_t[i]==barcode:
                    k=1
                    self.barcode_qty[i]+=1
                    break
                else:
                    k=0


            sql="SELECT 貨名,顏色,尺寸 FROM inventory where Barcode=?"
            try:

                temp=self.db.search_item(self.db.conn,sql,(barcode,))
                name=temp[0]
                color=temp[1]
                size=temp[2]
                row=self.ui.tableWidget.rowCount()
                self.ui.tableWidget.insertRow(row)
                self.ui.tableWidget.setItem(row,0,QTableWidgetItem(barcode))
                self.ui.tableWidget.setItem(row,1,QTableWidgetItem(name))
                self.ui.tableWidget.setItem(row,2,QTableWidgetItem(color))
                self.ui.tableWidget.setItem(row,3,QTableWidgetItem(size))
                if k==0:
                    self.barcode_t.append(barcode)
                    self.barcode_qty.append(1)

                print(self.barcode_t)
                print(self.barcode_qty)
            except TypeError as e:
                print("==========",e,"==========")
                msg=QMessageBox()
                msg.setWindowTitle("Information")
                msg.setText("物品不存在!")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()



            self.ui.lineEdit_2.clear()



if __name__=="__main__":
    app = QApplication(sys.argv)
    app.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling, True)
    w=out_show()
    sys.exit(app.exec_())
