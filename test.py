from openpyxl import load_workbook
import sqlite3
#new_wb=Workbook()

conn=sqlite3.connect("test.db")
cursor=conn.cursor()
wb=load_workbook(filename="US存貨.xlsx")
ws=wb["工作表1"]
cursor.execute("CREATE TABLE IF NOT EXISTS test(id integer);")



for j in range(1,ws.max_column):
    name=ws.cell(row=1,column=j).value
    add_column="ALTER TABLE test ADD COLUMN "+ name + ";"
    #conn.execute(add_column)
sql="INSERT INTO test (rowid,id,編號,更新日期, 產品編號, 產品名, 產品類別, 產品型號, 數量, 成本, 售價) values(?,?,?,?,?,?,?,?,?,?,?)"
for i in range(2,ws.max_row):
    row=[]
    row.append(i)
    for j in range(1,ws.max_column):
        data=ws.cell(row=i,column=j).value
        row.append(data)

    cursor.execute(sql,row)
conn.commit()
cursor.execute("SELECT * from test")
rows=cursor.fetchall()
print(rows)
for row in rows:
    print(row)

print("finish")
conn.close()
